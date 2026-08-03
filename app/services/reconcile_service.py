"""
Warehouse reconciliation: compare "my list" (Google Sheet export) against the
prep-center warehouse export.

Primary key:   (tracking_number, ASIN)  with quantity comparison.
Secondary key: ASIN only — for prep rows that have no tracking number
               (e.g. Inbound = "return") and my rows without a tracking.

The prep export has ONE ROW PER PHYSICAL UNIT (no quantity column) — quantity of
a (track, ASIN) group is the number of rows. Each row also carries a per-unit
"Cost USD", so the group's cost is the sum across its rows.

Both files are read directly from the uploaded bytes; nothing touches the DB.
"""

from __future__ import annotations

import csv
import io
from collections import defaultdict
from dataclasses import dataclass, field
from typing import Optional

# Inbound / origin values that are NOT tracking numbers → ASIN-only bucket.
_NON_TRACKING = {"return", "returns", "повернення", "damaged", "lost", "n/a", "-", "—", ""}
# Cell values that mean "no ASIN here".
_EMPTY_ASIN = {"", "-", "—", "n/a", "none"}


# ── File loading ──────────────────────────────────────────────────────────────

def _load_rows(data: bytes, filename: str) -> list[list]:
    """Read an .xlsx or .csv upload into a list of rows (list of cell values)."""
    name = (filename or "").lower()
    if name.endswith(".csv"):
        text = data.decode("utf-8-sig", errors="replace")
        return [list(r) for r in csv.reader(io.StringIO(text))]

    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    return [list(r) for r in ws.iter_rows(values_only=True)]


def _cells(header: list) -> list[str]:
    return [(str(c).strip().lower() if c is not None else "") for c in header]


def _find_col(header: list, *candidate_groups: list[str]) -> Optional[int]:
    """First header index matching a candidate, honouring group priority."""
    cells = _cells(header)
    for group in candidate_groups:
        for cand in group:
            for i, cell in enumerate(cells):
                if cand in cell:
                    return i
    return None


def _find_cols(header: list, candidates: list[str], exclude: list[str] | None = None) -> list[int]:
    """
    Ordered list of unique column indices matching the candidates, in candidate
    priority order. Columns whose header contains an `exclude` term are skipped.
    Used for ASIN, where the preferred column may be blank on a given row.
    """
    cells = _cells(header)
    exclude = exclude or []
    out: list[int] = []
    for cand in candidates:
        for i, cell in enumerate(cells):
            if cand in cell and i not in out and not any(x in cell for x in exclude):
                out.append(i)
    return out


def _as_int(val, default: int = 1) -> int:
    if val is None or val == "":
        return default
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default


def _as_float(val) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(str(val).replace("$", "").replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _norm_track(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip().upper()
    if s.lower() in _NON_TRACKING:
        return None
    return s


def _norm_asin(val) -> str:
    if val is None:
        return ""
    s = str(val).strip().upper()
    return "" if s.lower() in _EMPTY_ASIN else s


# ── Parsing each file into normalised entries ─────────────────────────────────

@dataclass
class _Parsed:
    tracked_qty: dict = field(default_factory=lambda: defaultdict(int))     # (track, asin) -> qty
    tracked_cost: dict = field(default_factory=lambda: defaultdict(float))  # (track, asin) -> cost sum
    untracked_qty: dict = field(default_factory=lambda: defaultdict(int))   # asin -> qty
    untracked_cost: dict = field(default_factory=lambda: defaultdict(float))
    has_cost: dict = field(default_factory=lambda: defaultdict(bool))       # key -> any cost seen
    titles: dict = field(default_factory=dict)                              # asin -> title
    row_count: int = 0
    header_found: dict = field(default_factory=dict)


def _parse(data: bytes, filename: str, is_prep: bool) -> _Parsed:
    rows = _load_rows(data, filename)
    result = _Parsed()
    if not rows:
        return result

    header = rows[0]
    if is_prep:
        track_i = _find_col(header, ["inbound"], ["origin ref"], ["tracking", "track"])
        asin_idxs = _find_cols(
            header,
            ["actual asin", "display asin", "pricing asin", "sku asin", "asin"],
        )
        qty_i = _find_col(header, ["qty", "quantity"])   # usually absent → 1 per row
        cost_i = _find_col(header, ["cost usd"], ["cost price"], ["собівар"])
        title_i = _find_col(header, ["sku title"], ["title"], ["product"])
    else:
        track_i = _find_col(header, ["track", "трек", "inbound", "origin ref"])
        asin_idxs = _find_cols(header, ["asin", "асін"])
        qty_i = _find_col(header, ["qty", "quantity", "кільк", "count", "к-сть"])
        cost_i = _find_col(header, ["cost", "собівар", "собівартість"])
        title_i = _find_col(header, ["title", "назва", "product"])

    result.header_found = {
        "tracking": track_i is not None,
        "asin": bool(asin_idxs),
        "qty": qty_i is not None,
        "cost": cost_i is not None,
    }

    def cell(row, idx):
        return row[idx] if (idx is not None and idx < len(row)) else None

    for row in rows[1:]:
        if not any(c not in (None, "") for c in row):
            continue

        # ASIN: first non-empty across the priority columns (Actual is often "-")
        asin = ""
        for idx in asin_idxs:
            asin = _norm_asin(cell(row, idx))
            if asin:
                break

        track = _norm_track(cell(row, track_i))
        qty = _as_int(cell(row, qty_i), default=1)
        if qty <= 0:
            qty = 1
        cost = _as_float(cell(row, cost_i))

        if not asin and not track:
            continue
        result.row_count += 1

        if title_i is not None and asin and asin not in result.titles:
            t = cell(row, title_i)
            if t:
                result.titles[asin] = str(t).strip()

        # Cost is per-unit on both sides (prep = 1 row per unit; my file = per-1pc),
        # so the line total is cost × qty.
        if track:
            key = (track, asin)
            result.tracked_qty[key] += qty
            if cost is not None:
                result.tracked_cost[key] += cost * qty
                result.has_cost[key] = True
        else:
            result.untracked_qty[asin] += qty
            if cost is not None:
                result.untracked_cost[asin] += cost * qty
                result.has_cost[("", asin)] = True

    return result


# ── Reconciliation ────────────────────────────────────────────────────────────

def _status(mine: int, prep: int) -> str:
    if mine > 0 and prep > 0:
        return "match" if mine == prep else "qty_mismatch"
    if mine > 0 and prep == 0:
        return "missing_in_prep"   # I have it, warehouse doesn't
    return "missing_in_mine"       # warehouse has it, I don't


def _cost_or_none(has_side: bool, val: float) -> Optional[float]:
    return round(val, 2) if has_side else None


def reconcile(my_bytes: bytes, my_name: str,
              prep_bytes: bytes, prep_name: str) -> dict:
    mine = _parse(my_bytes, my_name, is_prep=False)
    prep = _parse(prep_bytes, prep_name, is_prep=True)

    titles = {**mine.titles, **prep.titles}

    # Primary: tracking + ASIN
    tracked_rows = []
    cost_rows = []
    for key in set(mine.tracked_qty) | set(prep.tracked_qty):
        track, asin = key
        m, p = mine.tracked_qty.get(key, 0), prep.tracked_qty.get(key, 0)
        status = _status(m, p)
        tracked_rows.append({
            "tracking": track, "asin": asin, "title": titles.get(asin, ""),
            "mine": m, "prep": p, "status": status,
        })
        # Cost comparison only where track + qty agree
        if status == "match":
            my_c = _cost_or_none(mine.has_cost.get(key, False), mine.tracked_cost.get(key, 0.0))
            pr_c = _cost_or_none(prep.has_cost.get(key, False), prep.tracked_cost.get(key, 0.0))
            diff = None
            if my_c is not None or pr_c is not None:
                diff = round((my_c or 0.0) - (pr_c or 0.0), 2)
            cost_rows.append({
                "tracking": track, "asin": asin, "title": titles.get(asin, ""),
                "qty": m, "my_cost": my_c, "prep_cost": pr_c, "diff": diff,
            })

    # Secondary: ASIN-only (no tracking on either side)
    untracked_rows = []
    for asin in set(mine.untracked_qty) | set(prep.untracked_qty):
        m, p = mine.untracked_qty.get(asin, 0), prep.untracked_qty.get(asin, 0)
        untracked_rows.append({
            "asin": asin, "title": titles.get(asin, ""),
            "mine": m, "prep": p, "status": _status(m, p),
        })

    order = {"missing_in_prep": 0, "missing_in_mine": 1, "qty_mismatch": 2, "match": 3}
    tracked_rows.sort(key=lambda r: (order[r["status"]], r["tracking"], r["asin"]))
    untracked_rows.sort(key=lambda r: (order[r["status"]], r["asin"]))
    # Cost table: biggest discrepancies first, then the exact matches
    cost_rows.sort(key=lambda r: (-abs(r["diff"]) if r["diff"] is not None else 0,
                                  r["tracking"], r["asin"]))

    summary = defaultdict(int)
    for r in tracked_rows + untracked_rows:
        summary[r["status"]] += 1
    summary["total"] = len(tracked_rows) + len(untracked_rows)
    summary["ok"] = summary["match"]
    summary["problems"] = summary["total"] - summary["match"]
    summary["cost_diff"] = sum(1 for r in cost_rows if r["diff"] not in (None, 0, 0.0))

    return {
        "tracked": tracked_rows,
        "untracked": untracked_rows,
        "cost": cost_rows,
        "summary": dict(summary),
        "meta": {
            "my_rows": mine.row_count,
            "prep_rows": prep.row_count,
            "my_headers": mine.header_found,
            "prep_headers": prep.header_found,
        },
    }


# ── Template file for the user's "my list" upload ─────────────────────────────

def build_template_xlsx() -> bytes:
    """Minimal xlsx: tracking / asin / qty / cost (cost per 1 unit)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My list"

    headers = ["tracking", "asin", "qty", "cost"]
    fill = PatternFill("solid", fgColor="4F46E5")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill

    # cost is PER 1 UNIT — the line total is cost × qty
    examples = [
        ["1Z14V49E0337961675", "B0BYFJD8XX", 1, 434.75],
        ["1Z7R65990303240359", "B09J99GVXX", 4, 57.28],
        ["", "B00V525TXX", 1, 18.18],  # no tracking (return) — matched by ASIN
    ]
    for r in examples:
        ws.append(r)

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
